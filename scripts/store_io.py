"""SQLite repositories for public opportunities and the private pipeline.

Operational source of truth:
  <project>/data/public-opportunities.db   public job facts
  <project>/private/pipeline.db            application/contact/status facts

CSV/XLSX files are explicit, atomic exports. A legacy ``data/recruiting.db`` is
imported once when the split stores are first opened.
"""

from __future__ import annotations

import csv
import json
import os
import sqlite3
import sys
import tempfile
from datetime import date
from pathlib import Path

from schema_defs import (JOB_LIST_COLUMNS, TRACKER_COLUMNS, STATUSES,
                         REMOTE_POLICIES, POSTING_STATUSES, OUTCOMES)

ROOT = Path(__file__).resolve().parent.parent
KEYS = {"job_list": "job_id", "tracker": "application_id"}
COLS = {"job_list": JOB_LIST_COLUMNS, "tracker": TRACKER_COLUMNS}
CLEAR = object()
SCHEMA_VERSION = 3


def project_dir() -> Path:
    env = os.environ.get("RECRUITING_PROJECT_DIR")
    if env:
        p = Path(env).resolve()
    else:
        ap = ROOT / "active-project.json"
        if not ap.exists():
            sys.exit("FAIL: no active project. Run rolescout init or set RECRUITING_PROJECT_DIR.")
        with open(ap, encoding="utf-8") as f:
            p = (ROOT / json.load(f).get("active", "")).resolve()
    if not (p / "project.json").exists():
        sys.exit(f"FAIL: {p} is not a search project (missing project.json).")
    return p


def public_db_path() -> Path:
    return project_dir() / "data" / "public-opportunities.db"


def private_db_path() -> Path:
    return project_dir() / "private" / "pipeline.db"


def legacy_db_path() -> Path:
    return project_dir() / "data" / "recruiting.db"


def db_path(store: str = "job_list") -> Path:
    return public_db_path() if store == "job_list" else private_db_path()


def xlsx_path(store: str = "tracker") -> Path:
    return (project_dir() / "private" / "exports" / "pipeline.xlsx"
            if store == "tracker" else project_dir() / "exports" / "public-opportunities.xlsx")


def mirrors() -> dict[str, Path]:
    return {"job_list": project_dir() / "exports" / "public-opportunities.csv",
            "tracker": project_dir() / "private" / "exports" / "pipeline.csv"}


def _enum(vals) -> str:
    return ", ".join(f"'{v}'" for v in sorted(vals))


PUBLIC_DDL = f"""
CREATE TABLE IF NOT EXISTS job_list (
  job_id TEXT PRIMARY KEY, captured_at TEXT NOT NULL, company TEXT NOT NULL,
  title TEXT NOT NULL, job_group TEXT DEFAULT '', location TEXT DEFAULT '',
  remote_policy TEXT DEFAULT '' CHECK (remote_policy IN ('', {_enum(REMOTE_POLICIES)})),
  source_url TEXT NOT NULL, job_page_url TEXT DEFAULT '',
  posting_status TEXT DEFAULT '' CHECK (posting_status IN ('', {_enum(POSTING_STATUSES)})),
  seniority TEXT DEFAULT '', must_have_requirements TEXT DEFAULT '',
  nice_to_have_requirements TEXT DEFAULT '', jd_summary TEXT DEFAULT '',
  fit_score TEXT DEFAULT '' CHECK (fit_score IN ('', '1', '2', '3', '4', '5')),
  priority TEXT DEFAULT '' CHECK (priority IN ('', 'high', 'medium', 'low')),
  notes TEXT DEFAULT '', last_seen_at TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS job_visibility (
  job_id TEXT PRIMARY KEY REFERENCES job_list(job_id) ON DELETE CASCADE,
  position INTEGER NOT NULL
);
CREATE TABLE IF NOT EXISTS job_visibility_meta (
  singleton INTEGER PRIMARY KEY CHECK(singleton=1), initialized INTEGER NOT NULL DEFAULT 0
);
INSERT OR IGNORE INTO job_visibility_meta(singleton, initialized) VALUES(1, 0);
CREATE TABLE IF NOT EXISTS job_source_state (
  job_id TEXT NOT NULL REFERENCES job_list(job_id) ON DELETE CASCADE,
  source_key TEXT NOT NULL,
  company TEXT DEFAULT '', provider TEXT DEFAULT '', source_url TEXT DEFAULT '',
  first_seen_at TEXT NOT NULL, last_seen_at TEXT NOT NULL,
  last_checked_at TEXT NOT NULL, last_seen_run_id TEXT DEFAULT '',
  source_status TEXT NOT NULL CHECK (source_status IN ('open', 'removed')),
  PRIMARY KEY(job_id, source_key)
);
CREATE INDEX IF NOT EXISTS idx_job_source_state_source
  ON job_source_state(source_key, source_status);
CREATE TABLE IF NOT EXISTS source_scan_state (
  source_key TEXT PRIMARY KEY, company TEXT DEFAULT '', provider TEXT DEFAULT '',
  source_url TEXT DEFAULT '', last_run_id TEXT DEFAULT '',
  last_checked_at TEXT NOT NULL, scan_status TEXT DEFAULT '',
  authoritative INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE IF NOT EXISTS export_meta (
  name TEXT PRIMARY KEY, revision INTEGER NOT NULL DEFAULT 0,
  exported_revision INTEGER NOT NULL DEFAULT 0, exported_at TEXT DEFAULT ''
);
INSERT OR IGNORE INTO export_meta(name, revision) VALUES('job_list', 0);
CREATE TABLE IF NOT EXISTS migration_state (
  name TEXT PRIMARY KEY, completed_at TEXT NOT NULL
);
"""

PRIVATE_DDL = f"""
CREATE TABLE IF NOT EXISTS private.tracker (
  application_id TEXT PRIMARY KEY, job_id TEXT NOT NULL, company TEXT NOT NULL,
  title TEXT NOT NULL, job_group TEXT DEFAULT '',
  status TEXT NOT NULL CHECK (status IN ({_enum(set(STATUSES))})),
  applied_at TEXT DEFAULT '', resume_version TEXT DEFAULT '',
  linkedin_version TEXT DEFAULT '', contact TEXT DEFAULT '', next_action TEXT DEFAULT '',
  next_action_due TEXT DEFAULT '', last_updated TEXT NOT NULL,
  outcome TEXT DEFAULT '' CHECK (outcome IN ('', {_enum(set(o for o in OUTCOMES if o))})),
  notes TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS private.export_meta (
  name TEXT PRIMARY KEY, revision INTEGER NOT NULL DEFAULT 0,
  exported_revision INTEGER NOT NULL DEFAULT 0, exported_at TEXT DEFAULT ''
);
INSERT OR IGNORE INTO private.export_meta(name, revision) VALUES('tracker', 0);
"""


def connect() -> sqlite3.Connection:
    public = public_db_path()
    private = private_db_path()
    public.parent.mkdir(parents=True, exist_ok=True)
    private.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(public)
    con.execute("PRAGMA journal_mode = MEMORY")
    con.execute("PRAGMA foreign_keys = ON")
    con.execute("ATTACH DATABASE ? AS private", (str(private),))
    con.executescript(PUBLIC_DDL)
    con.executescript(PRIVATE_DDL)
    con.execute(f"PRAGMA user_version = {SCHEMA_VERSION}")
    con.execute(f"PRAGMA private.user_version = {SCHEMA_VERSION}")
    _migrate_removed_posting_status(con)
    _import_legacy_once(con)
    con.commit()
    return con


def _migrate_removed_posting_status(con: sqlite3.Connection) -> None:
    row = con.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='job_list'").fetchone()
    if not row or "'removed'" in (row[0] or ""):
        return
    cols = ", ".join(JOB_LIST_COLUMNS)
    with con:
        con.execute("ALTER TABLE job_list RENAME TO job_list_old")
        con.executescript(PUBLIC_DDL)
        con.execute(f"INSERT INTO job_list ({cols}) SELECT {cols} FROM job_list_old")
        con.execute("DROP TABLE job_list_old")


def _import_legacy_once(con: sqlite3.Connection) -> None:
    legacy = legacy_db_path()
    if con.execute("SELECT 1 FROM migration_state WHERE name='legacy-combined-v1'").fetchone():
        return
    if not legacy.exists():
        con.execute("INSERT INTO migration_state(name, completed_at) VALUES(?, ?)",
                    ("legacy-combined-v1", date.today().isoformat()))
        return
    if con.execute("SELECT COUNT(*) FROM job_list").fetchone()[0] or con.execute(
            "SELECT COUNT(*) FROM private.tracker").fetchone()[0]:
        con.execute("INSERT INTO migration_state(name, completed_at) VALUES(?, ?)",
                    ("legacy-combined-v1", date.today().isoformat()))
        return
    old = sqlite3.connect(legacy)
    old.row_factory = sqlite3.Row
    try:
        tables = {row[0] for row in old.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        if "job_list" in tables:
            upsert("job_list", [dict(row) for row in old.execute("SELECT * FROM job_list")], con)
        if "tracker" in tables:
            upsert("tracker", [dict(row) for row in old.execute("SELECT * FROM tracker")], con)
    finally:
        old.close()
    con.execute("INSERT INTO migration_state(name, completed_at) VALUES(?, ?)",
                ("legacy-combined-v1", date.today().isoformat()))


def _table(store: str) -> str:
    if store not in COLS:
        raise ValueError(f"unknown store: {store}")
    return "job_list" if store == "job_list" else "private.tracker"


def read_rows(store: str, con: sqlite3.Connection | None = None) -> list[dict]:
    own = con is None
    con = con or connect()
    con.row_factory = sqlite3.Row
    rows = [dict(row) for row in con.execute(f"SELECT * FROM {_table(store)}")]
    if own:
        con.close()
    return rows


def read_visible_job_rows(con: sqlite3.Connection | None = None) -> list[dict]:
    own = con is None
    con = con or connect()
    con.row_factory = sqlite3.Row
    initialized = con.execute(
        "SELECT initialized FROM job_visibility_meta WHERE singleton=1").fetchone()[0]
    query = ("SELECT j.* FROM job_visibility v JOIN job_list j ON j.job_id=v.job_id "
             "ORDER BY v.position" if initialized else "SELECT * FROM job_list")
    rows = [dict(row) for row in con.execute(query)]
    if own:
        con.close()
    return rows


def replace_visible_job_ids(job_ids: list[str], con: sqlite3.Connection | None = None) -> None:
    own = con is None
    con = con or connect()
    with con:
        con.execute("DELETE FROM job_visibility")
        con.executemany("INSERT INTO job_visibility(job_id, position) VALUES(?, ?)",
                        [(job_id, index) for index, job_id in enumerate(job_ids)])
        con.execute("UPDATE job_visibility_meta SET initialized=1 WHERE singleton=1")
    if own:
        con.close()


def read_project_rows(project: Path, store: str) -> list[dict]:
    """Read one store for an explicit project without active-project/env state."""
    path = (project / "data" / "public-opportunities.db" if store == "job_list"
            else project / "private" / "pipeline.db")
    legacy = project / "data" / "recruiting.db"
    if not path.exists() and legacy.exists():
        path = legacy
    if not path.exists():
        return []
    con = sqlite3.connect(path)
    con.row_factory = sqlite3.Row
    try:
        return [dict(row) for row in con.execute(f"SELECT * FROM {store}")]
    except sqlite3.Error:
        return []
    finally:
        con.close()


def _patch_value(value):
    if isinstance(value, dict) and value == {"$clear": True}:
        return CLEAR
    if isinstance(value, dict) and set(value) == {"$set"}:
        return str(value["$set"]).strip()
    return str(value).strip()


def upsert(store: str, rows: list[dict], con: sqlite3.Connection) -> tuple[int, int, int]:
    """Apply explicit patches and return inserted, changed, unchanged counts."""
    cols, key, table = COLS[store], KEYS[store], _table(store)
    inserted = changed = unchanged = 0
    for row in rows:
        existing_row = con.execute(f"SELECT * FROM {table} WHERE {key} = ?",
                                   (row.get(key),)).fetchone()
        if existing_row:
            current = dict(zip(cols, existing_row))
            sets: dict[str, str] = {}
            for col in cols:
                if col == key or col not in row:
                    continue
                if store == "job_list" and col == "captured_at":
                    continue  # immutable first-seen date; refresh last_seen_at instead
                value = _patch_value(row[col])
                if value is CLEAR:
                    value = ""
                elif value == "":
                    continue  # omitted/empty preserves; use {$clear:true} to clear
                if value != str(current.get(col, "")):
                    sets[col] = value
            if sets:
                assign = ", ".join(f"{col} = ?" for col in sets)
                con.execute(f"UPDATE {table} SET {assign} WHERE {key} = ?",
                            [*sets.values(), row[key]])
                changed += 1
            else:
                unchanged += 1
        else:
            vals = []
            for col in cols:
                value = _patch_value(row.get(col, ""))
                vals.append("" if value is CLEAR else value)
            con.execute(f"INSERT INTO {table} ({','.join(cols)}) VALUES ({','.join('?' * len(cols))})",
                        vals)
            inserted += 1
    if inserted or changed:
        meta = "export_meta" if store == "job_list" else "private.export_meta"
        con.execute(f"UPDATE {meta} SET revision = revision + 1 WHERE name = ?", (store,))
    return inserted, changed, unchanged


def revision(store: str, con: sqlite3.Connection | None = None) -> int:
    own = con is None
    con = con or connect()
    meta = "export_meta" if store == "job_list" else "private.export_meta"
    value = con.execute(f"SELECT revision FROM {meta} WHERE name = ?", (store,)).fetchone()[0]
    if own:
        con.close()
    return int(value)


def _atomic_csv(path: Path, store: str, rows: list[dict], rev: int) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=path.name + ".", suffix=".tmp", dir=path.parent)
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        with tmp.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=COLS[store], lineterminator="\n")
            writer.writeheader()
            writer.writerows(rows)
        os.replace(tmp, path)
        manifest = path.with_suffix(path.suffix + ".manifest.json")
        manifest.write_text(json.dumps({
            "schema": "rolescout-export-manifest-v1", "sensitivity": (
                "public" if store == "job_list" else "private"),
            "store": store, "database_revision": rev, "generated_at": date.today().isoformat(),
        }, indent=2) + "\n", encoding="utf-8")
    finally:
        tmp.unlink(missing_ok=True)


def export_views(con: sqlite3.Connection, *, public: bool = True,
                 private: bool = False, xlsx: bool = False) -> list[Path]:
    """Create explicit sensitivity-separated exports with atomic CSV replacement."""
    selected = (["job_list"] if public else []) + (["tracker"] if private else [])
    paths: list[Path] = []
    for store in selected:
        path = mirrors()[store]
        _atomic_csv(path, store, read_rows(store, con), revision(store, con))
        meta = "export_meta" if store == "job_list" else "private.export_meta"
        con.execute(f"UPDATE {meta} SET exported_revision=revision, exported_at=? WHERE name=?",
                    (date.today().isoformat(), store))
        paths.append(path)
    if xlsx:
        for store in selected:
            _export_xlsx(con, store)
            paths.append(xlsx_path(store))
    con.commit()
    return paths


def _export_xlsx(con: sqlite3.Connection, store: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = store
    cols = COLS[store]
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in read_rows(store, con):
        ws.append([row.get(col, "") for col in cols])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(col) + 2, 12), 40)
    path = xlsx_path(store)
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=path.stem + ".", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        wb.save(tmp)
        os.replace(tmp, path)
    finally:
        tmp.unlink(missing_ok=True)
